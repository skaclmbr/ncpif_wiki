# Used to build xml for import to the NCPIF Conservation Connections Wiki
#This is an initial test
# Scott Anderson Apr 15, 2020
# NC Wildlife Resources Commission
# scott.anderson@ncwildlife.org
# python 3.8

###################################################################
## End idea: provide a tool to convert wiki data to table format
## 	-Entities
## 	-Properties
## 	-Build and Parse Wiki XML imports/exports

###################################################################
## WIKITEXT STRUCTURE
## There are three main components to the wikitext on a page:
## 	1. Intro Template (includes list of linked items, formatted properties)
##  2. Supplementary information (description, etc.)
##  3. File link (if applicable)
##  4. Category declaration(s)
##  5. Property declaration(s)


###################################################################
## EXCEL
## results are stored in an excel workbook with the following structure:
## - entities (type, title)
## - categories (title, category)
## - properties (title, property, value)
## - wikitext (title, wikitext)


import xml.dom.minidom
from openpyxl import Workbook #allows connecting to databases
from openpyxl import load_workbook #allows connecting to databases
import datetime
from xml.sax.saxutils import escape

def openNewOutFile(ts, n):
	o_fn = str(ts.replace("-","").replace(":",""))+'_wiki_upload'+str(n)+'.xml'
	o = open(o_fn, 'w', encoding="utf-8")
	return o

def main():

	now = datetime.datetime.utcnow()
	timestamp = str(now.strftime('%Y-%m-%d'))+'T'+ str(now.strftime('%X'))+'Z' #run date - to be used in file name
	print (timestamp)

	mediawiki_open = '<mediawiki xmlns="http://www.mediawiki.org/xml/export-0.10/" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xsi:schemaLocation="http://www.mediawiki.org/xml/export-0.10/ http://www.mediawiki.org/xml/export-0.10.xsd" version="0.10" xml:lang="en">'
	siteinfo = '<siteinfo><sitename>NC Bird Conservation</sitename><dbname>ncbirdin_conwiki</dbname><base>http://wiki.ncpif.org/index.php?title=Main_Page</base><generator>MediaWiki 1.34.0</generator><case>first-letter</case><namespaces><namespace key="-2" case="first-letter">Media</namespace><namespace key="-1" case="first-letter">Special</namespace><namespace key="0" case="first-letter" /><namespace key="1" case="first-letter">Talk</namespace><namespace key="2" case="first-letter">User</namespace><namespace key="3" case="first-letter">User talk</namespace><namespace key="4" case="first-letter">NC Bird Conservation</namespace><namespace key="5" case="first-letter">NC Bird Conservation talk</namespace><namespace key="6" case="first-letter">File</namespace><namespace key="7" case="first-letter">File talk</namespace><namespace key="8" case="first-letter">MediaWiki</namespace><namespace key="9" case="first-letter">MediaWiki talk</namespace><namespace key="10" case="first-letter">Template</namespace><namespace key="11" case="first-letter">Template talk</namespace><namespace key="12" case="first-letter">Help</namespace><namespace key="13" case="first-letter">Help talk</namespace><namespace key="14" case="first-letter">Category</namespace><namespace key="15" case="first-letter">Category talk</namespace><namespace key="102" case="first-letter">Property</namespace><namespace key="103" case="first-letter">Property talk</namespace><namespace key="108" case="first-letter">Concept</namespace><namespace key="109" case="first-letter">Concept talk</namespace><namespace key="112" case="first-letter">smw/schema</namespace><namespace key="113" case="first-letter">smw/schema talk</namespace><namespace key="114" case="first-letter">Rule</namespace><namespace key="115" case="first-letter">Rule talk</namespace><namespace key="828" case="first-letter">Module</namespace><namespace key="829" case="first-letter">Module talk</namespace></namespaces></siteinfo>'
	mediawiki_close = '</mediawiki>'
	nl = '\n'
	propDelim = '|'

	# PRIMARY CATEGORIES
	types = ['Species','Habitat','Organization','Project','Plan','Geography','Priority'] #defines primary category for the entity

	# RECIPRICOL BEHAVIOR CHECK
	# takes the form of dictionary
	# enforces two-way links between some primary categories via properties
	# all categories have related property by adding "Has " to the front
	# is this needed?
	recipBehavior = {'Species':['Habitat','Organization','Project','Plan','Geography','Priority'],'Habitat':['Habitat','Organization','Project','Plan','Geography','Priority'],'Organization':['Habitat','Organization','Project','Plan','Geography','Priority'],'Project':['Habitat','Organization','Project','Plan','Geography','Priority'],'Plan':['Habitat','Organization','Project','Plan','Geography','Priority'],'Geography':['Habitat','Organization','Project','Plan','Geography','Priority'],'Priority':['Habitat','Organization','Project','Plan','Geography','Priority']}


	# DEFAULT COLUMNS
	colsEnt = {'type':'A','title':'B'}
	colsCat = {'title':'A','category':'B'}
	colsPro = {'title':'A','property':'B','value':'C'}
	colsWik = {'title':'A','wikitext':'B'}

	# ROW COUNTERS
	rowEnt = 1
	rowCat = 1
	rowPro = 1
	rowWik = 1


	###################################################################
	# EXCEL STRUCTURE
	wb = load_workbook(filename = '20200429173124_wiki_ncpif_for_upload.xlsx')

	wsEnt = wb['entities']
	wsCat = wb['categories']
	wsPro = wb['properties']
	wsWik = wb['wikitext']

	# ##############################################################
	# ## POPULATE DATA DICTIONARIES
	entities = {}
	categories = {}
	properties = {}
	wikitext = {}

	#################################################################
	# Get all entities
	rowEnt = 2
	currVal = wsEnt[colsEnt['title']+str(rowEnt)].value
	while currVal:
		# print (currVal)
		# entities.append(currVal)
		entities[currVal] = wsEnt[colsEnt['type']+str(rowEnt)].value
		categories[currVal] = []
		properties[currVal] = []
		wikitext[currVal] = ''
		rowEnt +=1
		currVal = wsEnt[colsEnt['title']+str(rowEnt)].value

	#################################################################
	# Get all category values
	rowCat = 2
	currEnt = wsCat[colsCat['title']+str(rowCat)].value
	while currEnt:
		currCat = wsCat[colsCat['category']+str(rowCat)].value

		# print (currEnt + " : " + currCat)
		categories[currEnt].append(currCat)

		rowCat +=1
		currEnt = wsCat[colsCat['title']+str(rowCat)].value

	# print (categories)

	#################################################################
	# Get all property values
	rowPro = 2
	currEnt = wsPro[colsPro['title']+str(rowPro)].value
	while currEnt:
		currProPro = wsPro[colsPro['property']+str(rowPro)].value
		currProVal = wsPro[colsPro['value']+str(rowPro)].value
		if currProVal:
			# currProVal = currProVal.replace("[[","").replace("]]","") #remove brackets from properties
			properties[currEnt].append([currProPro,currProVal])

		rowPro +=1
		currEnt = wsPro[colsPro['title']+str(rowPro)].value
	
	# print (properties)

	#################################################################
	# Get all wikitext values
	rowWik = 2
	currEnt = wsWik[colsWik['title']+str(rowWik)].value
	while currEnt:
		currWik = wsWik[colsWik['wikitext']+str(rowWik)].value

		# wikitext[currEnt].append(currWik)
		wikitext[currEnt]=currWik

		rowWik +=1
		currEnt = wsWik[colsWik['title']+str(rowWik)].value

	# print (wikitext)

	#################################################################
	## CHECK RECIPROCITY

	# Open missing page list document
	# OPEN OUTPUT FILE
	missing_fn = str(timestamp.replace("-","").replace(":",""))+'wiki_missing_pages.csv'
	missing = open(missing_fn, 'w', encoding="utf-8")
	missing.write(",".join(['page','ref_entity','ref_property'])+nl)

	allrecs_fn = str(timestamp.replace("-","").replace(":",""))+'wiki_allrecs.xml'
	allrecs = open(allrecs_fn, 'w', encoding="utf-8")
	allrecs.write(mediawiki_open+siteinfo)

	#Build properties to watch array
	propWatch = []
	for c in types:	propWatch.append("Has "+c)

	#loop through existing enities
	recipPro = properties #this is a copy of the properties dictionary for looping through
	recipEnt = entities #lookup array for entities (without messing with loop here)
	for e,t in entities.items():
		# print (nl+"=="+t+":"+e+"==")
		#loop through entities, check for property reciprocity
		recipLookupPro = "Has " + t #set the property that looks back towards this entity
		for p in recipPro[e]: #loop through properties of current entity
			
			targetEntity = p[1]
			if p[0] in propWatch: #is it a property linking primary categories?
				# print(p)
				bLinkFound = False #boolean indicating if property link exists
				
				# print(nl+"PRIMARY CATEGORY LINK TO: "+targetEntity)
				
				#check to make sure target exists!
				if targetEntity in recipEnt:
					# if targetEntity=='Estuarine Wetland Communities':
					# 	for x in properties[targetEntity]:print(x)
					# 	for x in recipPro[targetEntity]:print(x)
					for tP in properties[targetEntity]: #loop through property tuples of target
						# print (targetEntity + ": "+str(tP))
						if tP==[recipLookupPro,e]: #does the property match?
							# print ("--PROPERTY MATCH--"+nl)
							bLinkFound = True
							break #break out of peroperty values of target tuple

					if not bLinkFound: #not found, create property in target entity
						# print ("!!!PROPERTY NOT FOUND!!!")
						# print ("Add Has "+t+"="+e+" to " + targetEntity)
						properties[targetEntity].append(["Has "+t,e]) #add property to target
				else: #entity does not exist, add to missing page list
					missing.write(",".join([p[1],e,p[0]])+nl)


	#################################################################
	## BUILD WIKITEXT
	count = 1
	file_count = 1
	
	# OPEN OUTPUT FILE
	o = openNewOutFile(timestamp,file_count)
	# ADD LEADING INFO
	o.write(mediawiki_open+siteinfo)

	#Add Page Break Notation in Allrecs
	allrecs.write(nl+nl+"===PAGE START"+ str(file_count)+"========================================================================================================================="+nl)

	# LOOP THROUGH ENTITIES
	for e,t in entities.items():
		# print (nl+"=="+t+":"+e+"==")

		#build category string
		wtCat = ''
		for c in categories[e]:wtCat += "[[Category:"+c+"]]"
		# print (wtCat)
		#build properties string
		wtPro = ''
		pipe = ''
		for p in properties[e]:
			# print(p)
			# add to wikitext string
			wtPro += pipe+p[0]+"="+p[1]+nl
			if not pipe: pipe='|'

		# page_xml = f'{nl}<page>{nl}    <title>{e}</title>{nl}    <revision><id>338</id><parentid>333</parentid><timestamp>{t}</timestamp><contributor><username>Ncpif admin</username><id>1</id></contributor><model>wikitext</model><format>text/x-wiki</format>{nl}    <text xml:space="preserve">{nl}{wikitext[e]}{nl}{wtCat}{nl}{{{{#set:{nl}{wtPro}}}}}{nl}    </text>{nl}</revision>{nl}</page>{nl}'
		# page_xml = f'{nl}<page>{nl}    <title>{e}</title>{nl}    <revision><timestamp>{t}</timestamp><contributor><username>Ncpif admin</username><id>1</id></contributor><model>wikitext</model><format>text/x-wiki</format>{nl}    <text xml:space="preserve">{nl}{wikitext[e]}{nl}{wtCat}{nl}{{{{#set:{nl}{wtPro}}}}}{nl}    </text>{nl}</revision>{nl}</page>{nl}'
		page_contents = escape(f'{nl}{wikitext[e]}{nl}{wtCat}{nl}{{{{#set:{nl}{wtPro}}}}}{nl}')
		page_xml = f'{nl}<page>{nl}    <title>{e}</title>{nl}    <revision><timestamp>{t}</timestamp><contributor><username>Ncpif admin</username><id>1</id></contributor><model>wikitext</model><format>text/x-wiki</format>{nl}<text xml:space="preserve">{page_contents}</text>{nl}</revision>{nl}</page>{nl}'

		o.write(page_xml)
		allrecs.write(page_xml) #for troubleshooting
		count +=1
		if count % 50 ==0:
			#make a new file every 50 pages, mediawiki can't handle importing large files
			file_count+=1
			#Close out old file
			o.write(mediawiki_close)
			o.close()
			print (o.name + ' file closed'+nl)

			# Open new file
			o = openNewOutFile(timestamp,file_count)
			o.write(mediawiki_open+siteinfo)
			print (o.name + ' file opened'+nl)

			#Add Page Break Notation in Allrecs
			allrecs.write(nl+nl+"===PAGE START"+ str(file_count)+"========================================================================================================================="+nl)


		# print(page_xml)



	## FINISH UP PAGE
	allrecs.write(mediawiki_close) #for troubleshooting	
	count +=1
	# if count>3:
	# 	break #TESTING

	print(str(count) + " Pages Found")


if __name__=="__main__":
=======
# Used to build xml for import to the NCPIF Conservation Connections Wiki
#This is an initial test
# Scott Anderson Apr 15, 2020
# NC Wildlife Resources Commission
# scott.anderson@ncwildlife.org
# python 3.8

###################################################################
## End idea: provide a tool to convert wiki data to table format
## 	-Entities
## 	-Properties
## 	-Build and Parse Wiki XML imports/exports

###################################################################
## WIKITEXT STRUCTURE
## There are three main components to the wikitext on a page:
## 	1. Intro Template (includes list of linked items, formatted properties)
##  2. Supplementary information (description, etc.)
##  3. File link (if applicable)
##  4. Category declaration(s)
##  5. Property declaration(s)


###################################################################
## EXCEL
## results are stored in an excel workbook with the following structure:
## - entities (type, title)
## - categories (title, category)
## - properties (title, property, value)
## - wikitext (title, wikitext)


import xml.dom.minidom
from openpyxl import Workbook #allows connecting to databases
from openpyxl import load_workbook #allows connecting to databases
import datetime
from xml.sax.saxutils import escape

def openNewOutFile(ts, n):
	o_fn = str(ts.replace("-","").replace(":",""))+'_wiki_upload'+str(n)+'.xml'
	o = open('output_files/' + o_fn, 'w', encoding="utf-8")
	return o

def main():

	now = datetime.datetime.utcnow()
	timestamp = str(now.strftime('%Y-%m-%d'))+'T'+ str(now.strftime('%X'))+'Z' #run date - to be used in file name
	print (timestamp)

	mediawiki_open = '<mediawiki xmlns="http://www.mediawiki.org/xml/export-0.10/" xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xsi:schemaLocation="http://www.mediawiki.org/xml/export-0.10/ http://www.mediawiki.org/xml/export-0.10.xsd" version="0.10" xml:lang="en">'
	siteinfo = '<siteinfo><sitename>NC Bird Conservation</sitename><dbname>ncbirdin_conwiki</dbname><base>http://wiki.ncpif.org/index.php?title=Main_Page</base><generator>MediaWiki 1.34.0</generator><case>first-letter</case><namespaces><namespace key="-2" case="first-letter">Media</namespace><namespace key="-1" case="first-letter">Special</namespace><namespace key="0" case="first-letter" /><namespace key="1" case="first-letter">Talk</namespace><namespace key="2" case="first-letter">User</namespace><namespace key="3" case="first-letter">User talk</namespace><namespace key="4" case="first-letter">NC Bird Conservation</namespace><namespace key="5" case="first-letter">NC Bird Conservation talk</namespace><namespace key="6" case="first-letter">File</namespace><namespace key="7" case="first-letter">File talk</namespace><namespace key="8" case="first-letter">MediaWiki</namespace><namespace key="9" case="first-letter">MediaWiki talk</namespace><namespace key="10" case="first-letter">Template</namespace><namespace key="11" case="first-letter">Template talk</namespace><namespace key="12" case="first-letter">Help</namespace><namespace key="13" case="first-letter">Help talk</namespace><namespace key="14" case="first-letter">Category</namespace><namespace key="15" case="first-letter">Category talk</namespace><namespace key="102" case="first-letter">Property</namespace><namespace key="103" case="first-letter">Property talk</namespace><namespace key="108" case="first-letter">Concept</namespace><namespace key="109" case="first-letter">Concept talk</namespace><namespace key="112" case="first-letter">smw/schema</namespace><namespace key="113" case="first-letter">smw/schema talk</namespace><namespace key="114" case="first-letter">Rule</namespace><namespace key="115" case="first-letter">Rule talk</namespace><namespace key="828" case="first-letter">Module</namespace><namespace key="829" case="first-letter">Module talk</namespace></namespaces></siteinfo>'
	mediawiki_close = '</mediawiki>'
	nl = '\n'
	propDelim = '|'

	# PRIMARY CATEGORIES
	types = ['Species','Habitat','Organization','Project','Plan','Geography','Priority'] #defines primary category for the entity

	# RECIPRICOL BEHAVIOR CHECK
	# takes the form of dictionary
	# enforces two-way links between some primary categories via properties
	# all categories have related property by adding "Has " to the front
	# is this needed?
	recipBehavior = {'Species':['Habitat','Organization','Project','Plan','Geography','Priority'],'Habitat':['Habitat','Organization','Project','Plan','Geography','Priority'],'Organization':['Habitat','Organization','Project','Plan','Geography','Priority'],'Project':['Habitat','Organization','Project','Plan','Geography','Priority'],'Plan':['Habitat','Organization','Project','Plan','Geography','Priority'],'Geography':['Habitat','Organization','Project','Plan','Geography','Priority'],'Priority':['Habitat','Organization','Project','Plan','Geography','Priority']}


	# DEFAULT COLUMNS
	colsEnt = {'type':'A','title':'B'}
	colsCat = {'title':'A','category':'B'}
	colsPro = {'title':'A','property':'B','value':'C'}
	colsWik = {'title':'A','wikitext':'B'}

	# ROW COUNTERS
	rowEnt = 1
	rowCat = 1
	rowPro = 1
	rowWik = 1


	###################################################################
	# EXCEL STRUCTURE
	wb = load_workbook(filename = '20200429173124_wiki_ncpif_for_upload.xlsx')

	wsEnt = wb['entities']
	wsCat = wb['categories']
	wsPro = wb['properties']
	wsWik = wb['wikitext']

	# ##############################################################
	# ## POPULATE DATA DICTIONARIES
	entities = {}
	categories = {}
	properties = {}
	wikitext = {}

	#################################################################
	# Get all entities
	rowEnt = 2
	currVal = wsEnt[colsEnt['title']+str(rowEnt)].value
	while currVal:
		# print (currVal)
		# entities.append(currVal)
		entities[currVal] = wsEnt[colsEnt['type']+str(rowEnt)].value
		categories[currVal] = []
		properties[currVal] = []
		wikitext[currVal] = ''
		rowEnt +=1
		currVal = wsEnt[colsEnt['title']+str(rowEnt)].value

	#################################################################
	# Get all category values
	rowCat = 2
	currEnt = wsCat[colsCat['title']+str(rowCat)].value
	while currEnt:
		currCat = wsCat[colsCat['category']+str(rowCat)].value

		# print (currEnt + " : " + currCat)
		categories[currEnt].append(currCat)

		rowCat +=1
		currEnt = wsCat[colsCat['title']+str(rowCat)].value

	# print (categories)

	#################################################################
	# Get all property values
	rowPro = 2
	currEnt = wsPro[colsPro['title']+str(rowPro)].value
	while currEnt:
		currProPro = wsPro[colsPro['property']+str(rowPro)].value
		currProVal = wsPro[colsPro['value']+str(rowPro)].value
		if currProVal:
			# currProVal = currProVal.replace("[[","").replace("]]","") #remove brackets from properties
			properties[currEnt].append([currProPro,currProVal])

		rowPro +=1
		currEnt = wsPro[colsPro['title']+str(rowPro)].value
	
	# print (properties)

	#################################################################
	# Get all wikitext values
	rowWik = 2
	currEnt = wsWik[colsWik['title']+str(rowWik)].value
	while currEnt:
		currWik = wsWik[colsWik['wikitext']+str(rowWik)].value

		# wikitext[currEnt].append(currWik)
		wikitext[currEnt]=currWik

		rowWik +=1
		currEnt = wsWik[colsWik['title']+str(rowWik)].value

	# print (wikitext)

	#################################################################
	## CHECK RECIPROCITY

	# Open missing page list document
	# OPEN OUTPUT FILE
	missing_fn = str(timestamp.replace("-","").replace(":",""))+'wiki_missing_pages.csv'
	missing = open(missing_fn, 'w', encoding="utf-8")
	missing.write(",".join(['page','ref_entity','ref_property'])+nl)

	allrecs_fn = str(timestamp.replace("-","").replace(":",""))+'wiki_allrecs.xml'
	allrecs = open(allrecs_fn, 'w', encoding="utf-8")
	allrecs.write(mediawiki_open+siteinfo)

	#Build properties to watch array
	propWatch = []
	for c in types:	propWatch.append("Has "+c)

	#loop through existing enities
	recipPro = properties #this is a copy of the properties dictionary for looping through
	recipEnt = entities #lookup array for entities (without messing with loop here)
	for e,t in entities.items():
		# print (nl+"=="+t+":"+e+"==")
		#loop through entities, check for property reciprocity
		recipLookupPro = "Has " + t #set the property that looks back towards this entity
		for p in recipPro[e]: #loop through properties of current entity
			
			targetEntity = p[1]
			if p[0] in propWatch: #is it a property linking primary categories?
				# print(p)
				bLinkFound = False #boolean indicating if property link exists
				
				# print(nl+"PRIMARY CATEGORY LINK TO: "+targetEntity)
				
				#check to make sure target exists!
				if targetEntity in recipEnt:
					# if targetEntity=='Estuarine Wetland Communities':
					# 	for x in properties[targetEntity]:print(x)
					# 	for x in recipPro[targetEntity]:print(x)
					for tP in properties[targetEntity]: #loop through property tuples of target
						# print (targetEntity + ": "+str(tP))
						if tP==[recipLookupPro,e]: #does the property match?
							# print ("--PROPERTY MATCH--"+nl)
							bLinkFound = True
							break #break out of peroperty values of target tuple

					if not bLinkFound: #not found, create property in target entity
						# print ("!!!PROPERTY NOT FOUND!!!")
						# print ("Add Has "+t+"="+e+" to " + targetEntity)
						properties[targetEntity].append(["Has "+t,e]) #add property to target
				else: #entity does not exist, add to missing page list
					missing.write(",".join([p[1],e,p[0]])+nl)


	#################################################################
	## BUILD WIKITEXT
	count = 1
	file_count = 1
	
	# OPEN OUTPUT FILE
	o = openNewOutFile(timestamp,file_count)
	# ADD LEADING INFO
	o.write(mediawiki_open+siteinfo)

	#Add Page Break Notation in Allrecs
	allrecs.write(nl+nl+"===PAGE START"+ str(file_count)+"========================================================================================================================="+nl)

	# LOOP THROUGH ENTITIES
	for e,t in entities.items():
		# print (nl+"=="+t+":"+e+"==")

		#build category string
		wtCat = ''
		for c in categories[e]:wtCat += "[[Category:"+c+"]]"
		# print (wtCat)
		#build properties string
		wtPro = ''
		pipe = ''
		for p in properties[e]:
			# print(p)
			# add to wikitext string
			wtPro += pipe+p[0]+"="+p[1]+nl
			if not pipe: pipe='|'

		# page_xml = f'{nl}<page>{nl}    <title>{e}</title>{nl}    <revision><id>338</id><parentid>333</parentid><timestamp>{t}</timestamp><contributor><username>Ncpif admin</username><id>1</id></contributor><model>wikitext</model><format>text/x-wiki</format>{nl}    <text xml:space="preserve">{nl}{wikitext[e]}{nl}{wtCat}{nl}{{{{#set:{nl}{wtPro}}}}}{nl}    </text>{nl}</revision>{nl}</page>{nl}'
		# page_xml = f'{nl}<page>{nl}    <title>{e}</title>{nl}    <revision><timestamp>{t}</timestamp><contributor><username>Ncpif admin</username><id>1</id></contributor><model>wikitext</model><format>text/x-wiki</format>{nl}    <text xml:space="preserve">{nl}{wikitext[e]}{nl}{wtCat}{nl}{{{{#set:{nl}{wtPro}}}}}{nl}    </text>{nl}</revision>{nl}</page>{nl}'
		page_contents = escape(f'{nl}{wikitext[e]}{nl}{wtCat}{nl}{{{{#set:{nl}{wtPro}}}}}{nl}')
		page_xml = f'{nl}<page>{nl}    <title>{e}</title>{nl}    <revision><timestamp>{t}</timestamp><contributor><username>Ncpif admin</username><id>1</id></contributor><model>wikitext</model><format>text/x-wiki</format>{nl}<text xml:space="preserve">{page_contents}</text>{nl}</revision>{nl}</page>{nl}'

		o.write(page_xml)
		allrecs.write(page_xml) #for troubleshooting
		count +=1
		if count % 50 ==0:
			#make a new file every 50 pages, mediawiki can't handle importing large files
			file_count+=1
			#Close out old file
			o.write(mediawiki_close)
			o.close()
			print (o.name + ' file closed'+nl)

			# Open new file
			o = openNewOutFile(timestamp,file_count)
			o.write(mediawiki_open+siteinfo)
			print (o.name + ' file opened'+nl)

			#Add Page Break Notation in Allrecs
			allrecs.write(nl+nl+"===PAGE START"+ str(file_count)+"========================================================================================================================="+nl)


		# print(page_xml)



	## FINISH UP PAGE
	allrecs.write(mediawiki_close) #for troubleshooting	
	count +=1
	# if count>3:
	# 	break #TESTING

	print(str(count) + " Pages Found")


if __name__=="__main__":
	main();
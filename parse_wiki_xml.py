
# Used to process xml for the NCPIF Conservation Connections Wiki
#This is an initial test
# Scott Anderson Apr 15, 2020
# NC Wildlife Resources Commission
# scott.anderson@ncwildlife.org
# python 3.8

###################################################################
## Retrieving XML from wiki
#	1. Go to wiki.ncpif.org/index.php?title=Special:Export
#	2. Enter each category for inclusion in export file (e.g., 'Category:Species', 'Category:Species'). Make sure to include all primary entitites: 'Species','Habitat','Organization','Project','Plan','Geography','Priority'
#	3. Export file to folder with this python script
#	4. Change import_fn below to match export xml.


###################################################################
## End idea: provide a tool to convert wiki data to table format
## 	-Entities
## 	-Properties
## 	-Build and Parse Wiki XML imports/exports

###################################################################
## WIKITEXT STRUCTURE
## There are three main components to the wikitext on a page:
## 	1. Intro Template (includes list of linked items, formatted properties)
##  2. Supplementary information (description, etc.)
##  3. File link (if applicable)
##  4. Category declaration(s)
##  5. Property declaration(s)


###################################################################
## EXCEL
## results are stored in an excel workbook with the following structure:
## - entities (type, title)
## - categories (title, category)
## - properties (title, property, value)
## - wikitext (title, wikitext)



###################################################################
## QUESTIONS
## 	-What is the best storage/management database for input/output?
## 	-Should this be stored in MySQL online?

import xml.dom.minidom
from openpyxl import Workbook #allows connecting to databases
# from openpyxl import load_workbook #allows connecting to databases
from openpyxl.worksheet.table import Table, TableStyleInfo


def main():

	nl = '\n'
	propDelim = '|'

	# PRIMARY CATEGORIES
	types = ['Species','Habitat','Organization','Project','Plan','Geography','Priority'] #defines primary category for the entity

	# DEFAULT COLUMNS
	colsEnt = {'type':'A','title':'B'}
	colsCat = {'title':'A','category':'B'}
	colsPro = {'title':'A','property':'B','value':'C'}
	colsWik = {'title':'A','wikitext':'B'}

	# ROW COUNTERS
	rowEnt = 1
	rowCat = 1
	rowPro = 1
	rowWik = 1


	###################################################################
	# CREATE EXCEL FOR STORING DATA
	wb = Workbook()

	wsEnt = wb.active #get the first sheet
	wsEnt.title = "entities"
	wsCat = wb.create_sheet("categories")
	wsPro = wb.create_sheet("properties")
	wsWik = wb.create_sheet("wikitext")

	# Add headers
	wsEnt[colsEnt['type']+str(rowEnt)] = 'type'
	wsEnt[colsEnt['title']+str(rowEnt)] = 'title'
	rowEnt +=1

	wsCat[colsCat['title']+str(rowCat)] = 'title'
	wsCat[colsCat['category']+str(rowCat)] = 'category'
	rowCat +=1

	wsPro[colsPro['title']+str(rowPro)] = 'title'
	wsPro[colsPro['property']+str(rowPro)] = 'property'
	wsPro[colsPro['value']+str(rowPro)] = 'value'
	rowPro +=1

	wsWik[colsWik['title']+str(rowWik)] = 'title'
	wsWik[colsWik['wikitext']+str(rowWik)] = 'wikitext'
	rowWik +=1

	# LOAD XML FILE
	import_fn = "NC+Bird+Conservation-All-20200907141341.xml"
	timestamp = import_fn[25:39]

	doc = xml.dom.minidom.parse(import_fn)

	# LOOP THROUGH PAGES
	count = 1
	pages = doc.getElementsByTagName("page")
	print()

	for page in pages: #loop through all Wiki pages
		
		#################
		# TITLE - the key value for each page
		title = page.getElementsByTagName("title")[0].childNodes[0].nodeValue
		print (nl+"======" + str(title) + "======")

		#################
		# GET WIKITEXT
		wt = page.getElementsByTagName("text")[0].childNodes[0].nodeValue
		# wt = wt_raw.replace("{{#set:"+nl+"}}","")
		# print (wt)


		#################
		# REJECT TEMPLATES AND OTHER PAGES FOR NOW
		if "Category:" not in title:

			#################
			# CATEGORIES
			# loop to capture all categories
			cats = []
			priCat = ''
			startPos = wt.find("[[Category:")+11 #find first category
			currPos = startPos
			wtEnd = currPos-11 #define end of wikitext

			while currPos>=startPos:
				cEnd = wt.find("]]",currPos)
				currCat = wt[currPos:cEnd]

				if currCat in types: #check to see if one of main spp, add to entities sheet
					priCat = currCat
					wsEnt[colsEnt['type']+str(rowEnt)]=priCat
					wsEnt[colsEnt['title']+str(rowEnt)]=title
					rowEnt +=1

				wsCat[colsCat['title']+str(rowCat)]=title
				wsCat[colsCat['category']+str(rowCat)]=currCat
				rowCat +=1

				#advance cursor
				currPos = wt.find("[[Category:",cEnd)+11 #find next category

			cEnd +=2 # set end of category declaration

			#################
			# PROPERTIES
			# work with wikitext to extract properties
			# find #set declaration

			propDict = {}
			pStart = wt.find("{{#set:",cEnd)+7
			pEnd = wt.find("}}",pStart)
			
			if pEnd - pStart>5: #make sure there are some properties set
				propText = wt[pStart:pEnd]
				propList = propText.split(propDelim)
				for p in propList:
					itemValue=p.replace("\n","").split("=")
					print(itemValue)
					# print (itemValue)
					# print (colsPro['value']+str(rowPro))
					wsPro[colsPro['title']+str(rowPro)]=title
					wsPro[colsPro['property']+str(rowPro)]=itemValue[0]
					wsPro[colsPro['value']+str(rowPro)]=itemValue[1]
					rowPro +=1


			#################
			# WIKITEXT
			wsWik[colsWik['title']+str(rowWik)]=title
			wsWik[colsWik['wikitext']+str(rowWik)]=wt[:wtEnd] #get wikitext until categories start
			rowWik +=1

			# Includes Template, Supplementary Information, and File Links (#s 1-3 above)


		## FINISH UP PAGE
		count +=1
		# if count>3:
		# 	break #TESTING

	# SAVE EXCEL FILE
	# format/define tables
	style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,showLastColumn=False, showRowStripes=True, showColumnStripes=True)

	tabEnt = Table(displayName="entities",ref="A1:"+colsEnt['title']+str(rowEnt-1))
	tabEnt.tableStyleInfo = style
	wsEnt.add_table(tabEnt)
	tabCat = Table(displayName="categories",ref="A1:"+colsCat['category']+str(rowCat-1))
	tabCat.tableStyleInfo = style
	wsCat.add_table(tabCat)
	tabPro = Table(displayName="properties",ref="A1:"+colsPro['value']+str(rowPro-1))
	tabPro.tableStyleInfo = style
	wsPro.add_table(tabPro)
	tabWik = Table(displayName="wikitext",ref="A1:"+colsWik['wikitext']+str(rowWik-1))
	tabWik.tableStyleInfo = style
	wsWik.add_table(tabWik)



	wb.save(timestamp + '_wiki_ncpif.xlsx')

	print(str(count) + " Pages Found")


if __name__=="__main__":
=======
# Used to process xml for the NCPIF Conservation Connections Wiki
#This is an initial test
# Scott Anderson Apr 15, 2020
# NC Wildlife Resources Commission
# scott.anderson@ncwildlife.org
# python 3.8

###################################################################
## End idea: provide a tool to convert wiki data to table format
## 	-Entities
## 	-Properties
## 	-Build and Parse Wiki XML imports/exports

###################################################################
## WIKITEXT STRUCTURE
## There are three main components to the wikitext on a page:
## 	1. Intro Template (includes list of linked items, formatted properties)
##  2. Supplementary information (description, etc.)
##  3. File link (if applicable)
##  4. Category declaration(s)
##  5. Property declaration(s)


###################################################################
## EXCEL
## results are stored in an excel workbook with the following structure:
## - entities (type, title)
## - categories (title, category)
## - properties (title, property, value)
## - wikitext (title, wikitext)



###################################################################
## QUESTIONS
## 	-What is the best storage/management database for input/output?
## 	-Should this be stored in MySQL online?

import xml.dom.minidom
from openpyxl import Workbook #allows connecting to databases
# from openpyxl import load_workbook #allows connecting to databases
from openpyxl.worksheet.table import Table, TableStyleInfo


def main():

	nl = '\n'
	propDelim = '|'

	# PRIMARY CATEGORIES
	types = ['Species','Habitat','Organization','Project','Plan','Geography','Priority'] #defines primary category for the entity

	# DEFAULT COLUMNS
	colsEnt = {'type':'A','title':'B'}
	colsCat = {'title':'A','category':'B'}
	colsPro = {'title':'A','property':'B','value':'C'}
	colsWik = {'title':'A','wikitext':'B'}

	# ROW COUNTERS
	rowEnt = 1
	rowCat = 1
	rowPro = 1
	rowWik = 1


	###################################################################
	# CREATE EXCEL FOR STORING DATA
	wb = Workbook()

	wsEnt = wb.active #get the first sheet
	wsEnt.title = "entities"
	wsCat = wb.create_sheet("categories")
	wsPro = wb.create_sheet("properties")
	wsWik = wb.create_sheet("wikitext")

	# Add headers
	wsEnt[colsEnt['type']+str(rowEnt)] = 'type'
	wsEnt[colsEnt['title']+str(rowEnt)] = 'title'
	rowEnt +=1

	wsCat[colsCat['title']+str(rowCat)] = 'title'
	wsCat[colsCat['category']+str(rowCat)] = 'category'
	rowCat +=1

	wsPro[colsPro['title']+str(rowPro)] = 'title'
	wsPro[colsPro['property']+str(rowPro)] = 'property'
	wsPro[colsPro['value']+str(rowPro)] = 'value'
	rowPro +=1

	wsWik[colsWik['title']+str(rowWik)] = 'title'
	wsWik[colsWik['wikitext']+str(rowWik)] = 'wikitext'
	rowWik +=1

	# LOAD XML FILE
	import_fn = "NC+Bird+Conservation-All-20200429173124.xml"
	timestamp = import_fn[25:39]

	doc = xml.dom.minidom.parse(import_fn)

	# LOOP THROUGH PAGES
	count = 1
	pages = doc.getElementsByTagName("page")
	print()

	for page in pages: #loop through all Wiki pages
		
		#################
		# TITLE - the key value for each page
		title = page.getElementsByTagName("title")[0].childNodes[0].nodeValue
		print (nl+"======" + str(title) + "======")

		#################
		# GET WIKITEXT
		wt = page.getElementsByTagName("text")[0].childNodes[0].nodeValue
		# wt = wt_raw.replace("{{#set:"+nl+"}}","")
		# print (wt)


		#################
		# REJECT TEMPLATES AND OTHER PAGES FOR NOW
		if "Category:" not in title:

			#################
			# CATEGORIES
			# loop to capture all categories
			cats = []
			priCat = ''
			startPos = wt.find("[[Category:")+11 #find first category
			currPos = startPos
			wtEnd = currPos-11 #define end of wikitext

			while currPos>=startPos:
				cEnd = wt.find("]]",currPos)
				currCat = wt[currPos:cEnd]

				if currCat in types: #check to see if one of main spp, add to entities sheet
					priCat = currCat
					wsEnt[colsEnt['type']+str(rowEnt)]=priCat
					wsEnt[colsEnt['title']+str(rowEnt)]=title
					rowEnt +=1

				wsCat[colsCat['title']+str(rowCat)]=title
				wsCat[colsCat['category']+str(rowCat)]=currCat
				rowCat +=1

				#advance cursor
				currPos = wt.find("[[Category:",cEnd)+11 #find next category

			cEnd +=2 # set end of category declaration

			#################
			# PROPERTIES
			# work with wikitext to extract properties
			# find #set declaration

			propDict = {}
			pStart = wt.find("{{#set:",cEnd)+7
			pEnd = wt.find("}}",pStart)
			
			if pEnd - pStart>5: #make sure there are some properties set
				propText = wt[pStart:pEnd]
				propList = propText.split(propDelim)
				for p in propList:
					itemValue=p.replace("\n","").split("=")
					print(itemValue)
					# print (itemValue)
					# print (colsPro['value']+str(rowPro))
					wsPro[colsPro['title']+str(rowPro)]=title
					wsPro[colsPro['property']+str(rowPro)]=itemValue[0]
					wsPro[colsPro['value']+str(rowPro)]=itemValue[1]
					rowPro +=1


			#################
			# WIKITEXT
			wsWik[colsWik['title']+str(rowWik)]=title
			wsWik[colsWik['wikitext']+str(rowWik)]=wt[:wtEnd] #get wikitext until categories start
			rowWik +=1

			# Includes Template, Supplementary Information, and File Links (#s 1-3 above)


		## FINISH UP PAGE
		count +=1
		# if count>3:
		# 	break #TESTING

	# SAVE EXCEL FILE
	# format/define tables
	style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,showLastColumn=False, showRowStripes=True, showColumnStripes=True)

	tabEnt = Table(displayName="entities",ref="A1:"+colsEnt['title']+str(rowEnt-1))
	tabEnt.tableStyleInfo = style
	wsEnt.add_table(tabEnt)
	tabCat = Table(displayName="categories",ref="A1:"+colsCat['category']+str(rowCat-1))
	tabCat.tableStyleInfo = style
	wsCat.add_table(tabCat)
	tabPro = Table(displayName="properties",ref="A1:"+colsPro['value']+str(rowPro-1))
	tabPro.tableStyleInfo = style
	wsPro.add_table(tabPro)
	tabWik = Table(displayName="wikitext",ref="A1:"+colsWik['wikitext']+str(rowWik-1))
	tabWik.tableStyleInfo = style
	wsWik.add_table(tabWik)



	wb.save('output_files/' + timestamp + '_wiki_ncpif.xlsx')

	print(str(count) + " Pages Found")


if __name__=="__main__":
>>>>>>> 757f1a3192df6f426c4693325f63cb27a38cffdb
	main();